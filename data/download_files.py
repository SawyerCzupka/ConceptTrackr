import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import openpyxl
import os

path = "projects.xlsx"
# First download Excel spreadsheet off of GEF database:
# https://www.thegef.org/projects-operations/database --> scroll down left side --> export data in CSV format --> save in same folder as this
# IMPORTANT: save that file as an Excel spreadsheet instead of its original CSV (Comma Separated Values) format
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row
column = sheet_obj.max_column

url_base = "https://www.thegef.org/projects-operations/projects/"
# Base url for all GEF projects
print(url_base)

base_path = r"NEW"


def read():
    # Specific to Excel spreadsheet, change row/ column parameters to download files based off of country, region, etc.
    ls = []
    for i in range(1, row + 1):
        cell_obj = sheet_obj.cell(row=i, column=10)
        if cell_obj.value == 2015:  # year of interested projects (CHANGE)
            cell_obj1 = sheet_obj.cell(
                row=i, column=2
            )  # ID column from Excel spreadsheet
            ls.append(cell_obj1.value)
    return ls


def main():
    # list = read()
    list = [4751, 10255]
    for (
        i
    ) in (
        list
    ):  # Uses list just created of all the IDs for year specified, change to (1, 11304) to download ALL GEF projects
        url = url_base + str(i)
        response = requests.get(url)
        if response.status_code == 200:
            print(url + " Web site exists")
            response = requests.get(url)
            # Parse text
            soup = BeautifulSoup(response.text, "html.parser")
            # Find all hyperlinks on webpage
            links = soup.find_all("a")
            g = 0

            for link in links:
                try:
                    if ".pdf" in link.get("href", []):
                        g += 1
                        response = requests.get(link.get("href"))
                        pdf = open(
                            os.path.join(
                                base_path, "pdf" + str(i) + "_" + str(g) + ".pdf"
                            ),
                            "wb",
                        )  # Prints GEF pdfs as #ID_#file of project
                        pdf.write(response.content)
                        pdf.close()
                        # print(f"Link: {link}")
                        print("File ", g, " downloaded")
                except Exception as e:
                    print("pdf does not exist")
                    print(e)

        else:
            print("Web site does not exist")


if __name__ == "__main__":
    main()
